from datetime import datetime as dateuser
import string
import time
import json
import os
from openpyxl.utils import get_column_letter
import requests

def Twitch_Auth(config_file):
    # Get the credentials from the json config file
    absolute_path = os.path.abspath(config_file)
    with open(absolute_path, "r", encoding='utf-8') as json_file:
        data = json.load(json_file)
        client_id = data['client_id']
        client_secret = data['client_secret']

    body = {
        'client_id': client_id,
        'client_secret': client_secret,
        "grant_type": 'client_credentials'
    }
    r = requests.post('https://id.twitch.tv/oauth2/token', body)
    keys = r.json()
    headers = {
        'Client-ID': client_id,
        'Authorization': 'Bearer ' + keys['access_token']
    }
    stream = requests.get(
        'https://api.twitch.tv/helix/streams?first=100&language=fr', headers=headers)
    return stream


def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


def total_fr(stream_data, amount):
    total_fr = 0
    i = 0
    while i < amount:
        total_fr += stream_data['data'][i]['viewer_count']
        i += 1

    now = dateuser.now()
    today = now.strftime("%H:%M:%S")
    print(f"Last datas gathered at : {today}")
    return total_fr


def getline(username, ws):
    for columns in ws.iter_cols(1):
        for cell in columns:
            if cell.value == username:
                return(cell.column)
    return get_column_letter(len(ws['1']) + 1)
