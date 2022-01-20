import datetime
import os.path
import re
import sys
import time
import argparse

from datetime import datetime
from datetime import date

import openpyxl
from openpyxl.utils import get_column_letter
from func import Twitch_Auth, col2num, getline, total_fr

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='LsPDAGather')
    parser.add_argument('-f', '--file', help='config file', required=True)
    parser.add_argument('-o', '--output-dir', help='output directory for generated files', required=False, default='XLfiles')
    parser.add_argument('-r', '--refresh-rate', help='refresh rate in seconds', required=False, type=int, default=60)
    parser.add_argument('-n', '--number-of-streamers', help='number of streamers to gather', required=False, type=int, default=15)
    parser.add_argument('-i', '--iterations', help='number of requests to make', required=False, type=int, default=50)
    parser.add_argument('-a', '--data-amount', help='total number of streamers to generate data', type=int, required=False, default=98)
    args = parser.parse_args()    
    # Filename generation
    today = date.today()
    output_dir = args.output_dir
    try:
        os.makedirs(output_dir, exist_ok=True)
    except OSError as e:
        print("Creation of the directory %s failed: %s" % output_dir, e)

    filepath = f"StatTwitch{today}"
    cfilepath = f"StatTwitch{today}.xlsx"

    filepath = os.path.join(output_dir, filepath)
    cfilepath = os.path.join(output_dir, cfilepath)

    # Creates an excel file
    iter_path = 1
    while os.path.isfile(cfilepath):
        cfilepath = f"{filepath}_{iter_path:03d}.xlsx"
        iter_path += 1
    wb = openpyxl.Workbook()
    wb.guess_types = True
    ws = wb.active

    # Variables settings
    totalpda = float(0)
    i = 0
    t = 'B'
    col = 1
    firstpassage = True
    line = 1

    for n in range(0, args.iterations):
        totalviewers = 0

        # Getting datas from Twitch API
        stream = Twitch_Auth(args.file)
        stream_data = stream.json()
        stats_stream_fr = 0
        stats_stream_fr = total_fr(stream_data, args.data_amount)
        total_frbis = stats_stream_fr
        now = datetime.now()
        today = now.strftime("%H:%M:%S")
        i = 0
        totalpda = float(0)

        # Analizes the TOP_NB Twitch streamers
        for i in range(0, args.number_of_streamers + 1):
            pda = (stream_data['data'][i]['viewer_count'] * 100) / total_frbis
            pda = "{: .2f}".format(pda)
            username = stream_data['data'][i]['user_name']
            viewer_count = stream_data['data'][i]['viewer_count']
            title = stream_data['data'][i]['title']
            data = [username, pda]
            # Table format
            if firstpassage:
                ws[f'A{2}'] = today
                ws[f'{t}1'] = username
                ws[f'{t}2'] = float(pda)
            else:
                try:
                    t = get_column_letter(getline(username, ws))
                except Exception as e:
                    t = getline(username, ws)
                ws[f'A{line}'] = today
                ws[f"{t}1"] = username
                ws[f"{t}{line}"] = float(pda)
            totalpda = "{: .2f}".format(totalpda)
            totalpda = float(totalpda) + float(pda)
            t = col2num(t)
            t = get_column_letter(t + 1)
            i += 1
        line += 1
        firstpassage = False
        wb.save(cfilepath)
        print(f"[{n}/{args.iterations}]" + " Saving file: " + cfilepath)
        try:
            time.sleep(args.refresh_rate)
        except KeyboardInterrupt as e:
            print("Caught KeyboardInterrupt, terminating...")
            wb.save(cfilepath)
            print("Saving file: " + cfilepath)
            sys.exit(0)

